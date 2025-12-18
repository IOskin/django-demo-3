from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from shop.models import Category, Manufacturer, Supplier, Product


class Command(BaseCommand):
    help = "Импорт товаров из Excel файла Прил_2_ОЗ_Канцтовары-M1.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            default="Прил_2_ОЗ_Канцтовары-M1.xlsx",
            help="Путь к Excel-файлу (по умолчанию: Прил_2_ОЗ_Канцтовары-M1.xlsx)",
        )

    def handle(self, *args, **options):
        file_path = Path(options["path"])
        if not file_path.exists():
            raise CommandError(f"Файл не найден: {file_path}")

        wb = load_workbook(filename=str(file_path), data_only=True)

        def find_sheet_with_header(target_header: str):
            normalized_target = str(target_header).strip().lower()
            for ws in wb.worksheets:
                headers = [cell.value for cell in ws[1]]
                for header in headers:
                    if header is None:
                        continue
                    if str(header).strip().lower() == normalized_target:
                        return ws, headers
            available = {
                ws.title: [cell.value for cell in ws[1]] for ws in wb.worksheets
            }
            raise CommandError(
                f"Не удалось найти лист с колонкой заголовка {target_header!r}. "
                f"Найденные заголовки по листам: {available}"
            )

        ws, headers = find_sheet_with_header("Категория товара")

        def get_index(name: str) -> int:
            normalized_target = str(name).strip().lower()
            for idx, header in enumerate(headers):
                if header is None:
                    continue
                if str(header).strip().lower() == normalized_target:
                    return idx
            available = ", ".join(str(h) for h in headers)
            raise CommandError(
                f"В файле на листе {ws.title!r} нет колонки с заголовком: {name!r}. "
                f"Найденные заголовки: {available}"
            )

        idx_category = get_index("Категория товара")
        idx_name = get_index("Наименование товара")
        idx_manufacturer = get_index("Производитель")
        idx_supplier = get_index("Поставщик")
        idx_price = get_index("Цена")
        idx_unit = get_index("Единица измерения")
        idx_stock = get_index("Кол-во на складе")
        idx_discount = get_index("Действующая скидка")
        idx_description = get_index("Описание товара")
        idx_image = get_index("Фото")

        created_products = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[idx_name]:
                continue

            category_name = str(row[idx_category]).strip() if row[idx_category] else "Без категории"
            manufacturer_name = str(row[idx_manufacturer]).strip() if row[idx_manufacturer] else "Без производителя"
            supplier_name = str(row[idx_supplier]).strip() if row[idx_supplier] else "Без поставщика"

            category, _ = Category.objects.get_or_create(name=category_name)
            manufacturer, _ = Manufacturer.objects.get_or_create(name=manufacturer_name)
            supplier, _ = Supplier.objects.get_or_create(name=supplier_name)

            try:
                price = float(row[idx_price]) if row[idx_price] is not None else 0
            except (TypeError, ValueError):
                price = 0

            unit_raw = (str(row[idx_unit]).strip() if row[idx_unit] is not None else "").lower()
            unit_map = {
                "шт": "pcs",
                "шт.": "pcs",
                "уп": "pack",
                "уп.": "pack",
                "набор": "set",
            }
            unit = unit_map.get(unit_raw, "pcs")

            try:
                stock = int(row[idx_stock]) if row[idx_stock] is not None else 0
            except (TypeError, ValueError):
                stock = 0

            try:
                discount = int(row[idx_discount]) if row[idx_discount] is not None else 0
            except (TypeError, ValueError):
                discount = 0

            description = (
                str(row[idx_description]).strip() if row[idx_description] is not None else ""
            )
            image_value = row[idx_image]
            image_path = None
            if image_value:
                image_path = f"products/{str(image_value).strip()}"

            Product.objects.create(
                name=str(row[idx_name]).strip(),
                category=category,
                description=description,
                manufacturer=manufacturer,
                supplier=supplier,
                price=price,
                unit=unit,
                stock_quantity=stock,
                discount_percent=discount,
                image=image_path,
            )
            created_products += 1

        self.stdout.write(self.style.SUCCESS(f"Импорт завершён. Создано товаров: {created_products}"))

